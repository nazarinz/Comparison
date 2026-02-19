# ============================================================
# PGD Comparison Tracking — SAP vs Infor  |  PO Splitter 5.000 | Sales Analytics Dashboard
# ============================================================

import io
import sys
import re
import zipfile
from datetime import datetime, timedelta
from contextlib import nullcontext

import numpy as np
import pandas as pd
import streamlit as st

# ==== OpenPyXL (opsional, untuk ekspor Excel yang di-styling) ====
EXCEL_EXPORT_AVAILABLE = True
try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    EXCEL_EXPORT_AVAILABLE = False

# ================== Streamlit Config ==================
st.set_page_config(page_title="PGD Comparison & PO Splitter & Dashboard", layout="wide")
st.title("📦 PGD Comparison — SAP vs Infor  |  🧩 PO Splitter 5.000  |  📈 Sales Analytics")

# ================== Warna, Kolom, Format ==================
INFOR_COLOR  = "FFF9F16D"  # kuning lembut (header Infor)
RESULT_COLOR = "FFC6EFCE"  # hijau lembut (header Result_*)
OTHER_COLOR  = "FFD9D9D9"  # abu-abu muda (header lainnya)
DATE_FMT     = "m/d/yyyy"

INFOR_COLUMNS_FIXED = [
    "Order Status Infor","Infor Quantity","Infor Model Name","Infor Article No",
    "Infor Classification Code","Infor Delay/Early - Confirmation CRD",
    "Infor Delay - PO PSDD Update","Infor Lead time","Infor GPS Country",
    "Infor Ship-to Country","Infor FPD","Infor LPD","Infor CRD","Infor PSDD",
    "Infor PODD","Infor PD","Infor Delay - PO PD Update",
    "Infor Shipment Method",
    "Infor Market PO Number"  # <= baru
]

DELAY_EMPTY_COLUMNS = [
    "Delay/Early - Confirmation CRD",
    "Infor Delay/Early - Confirmation CRD",
    "Result_Delay_CRD",
    "Delay - PO PSDD Update",
    "Infor Delay - PO PSDD Update"
]

DATE_COLUMNS_PREF = [
    "Document Date","FPD","LPD","CRD","PSDD","FCR Date","PODD","PD","PO Date","Actual PGI",
    "Infor FPD","Infor LPD","Infor CRD","Infor PSDD","Infor PODD","Infor PD"
]

# ================== Helpers (Umum) ==================
def today_str_id():
    return (datetime.utcnow() + timedelta(hours=7)).strftime("%Y%m%d")

def status_ctx(label="Processing...", expanded=True):
    if hasattr(st, "status"):
        return st.status(label, expanded=expanded)
    st.info(label);  return nullcontext()

def _status_update(ctx, label=None, state=None):
    if hasattr(ctx, "update"):
        ctx.update(label=label, state=state)
    else:
        if state == "error": st.error(label or "")
        elif state == "complete": st.success(label or "")
        else: st.info(label or "")

@st.cache_data(show_spinner=False)
def read_excel_file(file):
    return pd.read_excel(file, engine="openpyxl")

@st.cache_data(show_spinner=False)
def read_csv_file(file):
    for enc in ("utf-8", "utf-8-sig", "latin1"):
        try:
            file.seek(0);  return pd.read_csv(file, encoding=enc)
        except Exception:
            continue
    file.seek(0);  return pd.read_csv(file)

def convert_date_columns(df):
    date_cols = [
        'Document Date','FPD','LPD','CRD','PSDD','FCR Date','PODD','PD','PO Date','Actual PGI',
        'Infor CRD','Infor PD','Infor PSDD','Infor FPD','Infor LPD','Infor PODD'
    ]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df

def load_sap(sap_df):
    df = sap_df.copy()
    if "Quanity" in df.columns and "Quantity" not in df.columns:
        df.rename(columns={'Quanity': 'Quantity'}, inplace=True)
    if "PO No.(Full)" in df.columns:
        df["PO No.(Full)"] = df["PO No.(Full)"].astype(str).str.strip()
    df = convert_date_columns(df)
    return df

def load_infor_from_many_csv(csv_dfs):
    data_list = []
    required_cols = ['PO Statistical Delivery Date (PSDD)','Customer Request Date (CRD)','Line Aggregator']
    for i, df in enumerate(csv_dfs, start=1):
        if all(col in df.columns for col in required_cols):
            data_list.append(df);  st.success(f"Dibaca ✅ CSV ke-{i} (kolom wajib lengkap)")
        else:
            miss = [c for c in required_cols if c not in df.columns]
            st.warning(f"CSV ke-{i} dilewati ⚠️ (kolom wajib hilang: {miss})")
    if not data_list: return pd.DataFrame()
    return pd.concat(data_list, ignore_index=True)

def process_infor(df_all):
    selected_columns = [
        'Order #','Order Status','Model Name','Article Number','Gps Customer Number',
        'Country/Region','Customer Request Date (CRD)','Plan Date','PO Statistical Delivery Date (PSDD)',
        'First Production Date','Last Production Date','PODD','Production Lead Time',
        'Class Code','Delay - Confirmation','Delay - PO Del Update','Quantity',
        'Delivery Delay Pd',
        'Shipment Method',
        'Market PO Number'  # <= baru
    ]
    missing_cols = [col for col in selected_columns if col not in df_all.columns]
    if missing_cols:
        st.error(f"Kolom hilang dalam data Infor: {missing_cols}")
        return pd.DataFrame()

    df_infor = df_all[selected_columns].copy()
    df_infor = df_infor.groupby('Order #', as_index=False).agg({
        'Order Status':'first','Model Name':'first','Article Number':'first','Gps Customer Number':'first',
        'Country/Region':'first','Customer Request Date (CRD)':'first','Plan Date':'first',
        'PO Statistical Delivery Date (PSDD)':'first','First Production Date':'first',
        'Last Production Date':'first','PODD':'first','Production Lead Time':'first',
        'Class Code':'first','Delay - Confirmation':'first','Delay - PO Del Update':'first',
        'Quantity':'sum',
        'Delivery Delay Pd':'first',
        'Shipment Method':'first',
        'Market PO Number':'first'  # <= baru
    })
    df_infor["Order #"] = df_infor["Order #"].astype(str).str.zfill(10).str.strip()

    rename_cols = {
        'Order Status':'Order Status Infor','Model Name':'Infor Model Name','Article Number':'Infor Article No',
        'Gps Customer Number':'Infor GPS Country','Country/Region':'Infor Ship-to Country',
        'Customer Request Date (CRD)':'Infor CRD','Plan Date':'Infor PD',
        'PO Statistical Delivery Date (PSDD)':'Infor PSDD','First Production Date':'Infor FPD',
        'Last Production Date':'Infor LPD','PODD':'Infor PODD','Production Lead Time':'Infor Lead time',
        'Class Code':'Infor Classification Code','Delay - Confirmation':'Infor Delay/Early - Confirmation CRD',
        'Delay - PO Del Update':'Infor Delay - PO PSDD Update','Quantity':'Infor Quantity',
        'Delivery Delay Pd':'Infor Delay - PO PD Update',
        'Shipment Method': 'Infor Shipment Method',
        'Market PO Number': 'Infor Market PO Number'  # <= baru
    }
    df_infor.rename(columns=rename_cols, inplace=True)
    st.info(f"Jumlah baris setelah proses Infor: {len(df_infor)}")
    return df_infor

def merge_sap_infor(df_sap, df_infor):
    df_sap = df_sap.copy();  df_infor = df_infor.copy()
    if 'PO No.(Full)' in df_sap.columns:
        df_sap['PO No.(Full)'] = df_sap['PO No.(Full)'].astype(str).str.zfill(10)
    if 'Order #' in df_infor.columns:
        df_infor['Order #'] = df_infor['Order #'].astype(str).str.zfill(10)
    return df_sap.merge(df_infor, how='left', left_on='PO No.(Full)', right_on='Order #')

# ====== Fix: Handle SAP 1 baris vs Infor multi-baris (hindari duplikat Quantity) ======
def match_qty_nearest(df_sap, df_infor, key="PO No.(Full)", qty_col="Quantity", infor_qty_col="Infor Quantity"):
    df_sap = df_sap.copy()
    df_sap["___sap_row_id"] = np.arange(len(df_sap))

    merged = df_sap.merge(df_infor, how="left", left_on=key, right_on="Order #")
    if merged.empty:
        return merged

    out_rows = []
    for po, group in merged.groupby(key, sort=False):
        n_sap_rows = group["___sap_row_id"].nunique() if "___sap_row_id" in group.columns else 1

        if n_sap_rows > 1:
            out_rows.extend(group.drop(columns=["___sap_row_id"], errors="ignore").to_dict("records"))
            continue

        if len(group) == 1:
            out_rows.append(group.drop(columns=["___sap_row_id"], errors="ignore").iloc[0].to_dict())
            continue

        sap_qty = pd.to_numeric(group.iloc[0].get(qty_col, 0), errors="coerce")
        diffs = (pd.to_numeric(group.get(infor_qty_col, 0), errors="coerce") - sap_qty).abs()
        idx_min = diffs.idxmin() if diffs.notna().any() else group.index[0]

        for idx, row in group.iterrows():
            row = row.copy()
            if idx != idx_min:
                row[qty_col] = np.nan
            out_rows.append(row.drop(labels=["___sap_row_id"], errors="ignore").to_dict())

    return pd.DataFrame(out_rows)

def fill_missing_dates(df):
    df = df.copy()
    df['Order Status Infor'] = df.get('Order Status Infor', pd.Series(dtype=str)).astype(str).str.strip().str.upper()
    for col in ['LPD','FPD','CRD','PD','PSDD','PODD']:
        if col not in df.columns: df[col] = pd.NaT
        df[col] = pd.to_datetime(df[col], errors='coerce')
    mask_open = df['Order Status Infor'].eq('OPEN')
    min_dates = df[['CRD','PD']].min(axis=1)
    df.loc[mask_open & df['LPD'].isna(),'LPD'] = min_dates
    df.loc[mask_open & df['FPD'].isna(),'FPD'] = min_dates
    df.loc[mask_open & df['PSDD'].isna(),'PSDD'] = df['CRD']
    df.loc[mask_open & df['PODD'].isna(),'PODD'] = df['CRD']
    return df

def clean_and_compare(df_merged):
    df_merged = df_merged.copy()

    for col in ["Quantity","Infor Quantity","Production Lead Time","Infor Lead time","Article Lead time"]:
        if col in df_merged.columns:
            df_merged[col] = pd.to_numeric(df_merged[col], errors="coerce").fillna(0).round(2)

    code_mapping = {
        '161':'01-0161','84':'03-0084','68':'02-0068','64':'04-0064','62':'02-0062','61':'01-0061',
        '51':'03-0051','46':'03-0046','7':'02-0007','3':'03-0003','2':'01-0002','1':'01-0001',
        '4':'04-0004','8':'02-0008','10':'04-0010','49':'03-0049','90':'04-0090','63':'03-0063','27':'04-0027'
    }
    def map_code_safely(x):
        try: return code_mapping.get(str(int(float(x))), x)
        except (ValueError, TypeError): return x

    if "Infor Delay/Early - Confirmation CRD" in df_merged.columns:
        df_merged["Infor Delay/Early - Confirmation CRD"] = (
            df_merged["Infor Delay/Early - Confirmation CRD"]
            .replace(['--','N/A','NULL'], pd.NA).apply(map_code_safely)
        )
    if "Infor Delay - PO PSDD Update" in df_merged.columns:
        df_merged["Infor Delay - PO PSDD Update"] = (
            df_merged["Infor Delay - PO PSDD Update"]
            .replace(['--','N/A','NULL'], pd.NA).apply(map_code_safely)
        )
    if "Infor Delay - PO PD Update" in df_merged.columns:
        df_merged["Infor Delay - PO PD Update"] = (
            df_merged["Infor Delay - PO PD Update"]
            .replace(['--','N/A','NULL'], pd.NA).apply(map_code_safely)
        )

    string_cols = [
        "Model Name","Infor Model Name","Article No","Infor Article No",
        "Classification Code","Infor Classification Code",
        "Ship-to Country","Infor Ship-to Country",
        "Ship-to-Sort1","Infor GPS Country",
        "Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
        "Delay - PO PSDD Update","Infor Delay - PO PSDD Update",
        "Delay - PO PD Update","Infor Delay - PO PD Update",
        "Infor Shipment Method"
    ]
    for col in string_cols:
        if col in df_merged.columns:
            df_merged[col] = df_merged[col].astype(str).str.strip().str.upper()
    if "Ship-to-Sort1" in df_merged.columns:
        df_merged["Ship-to-Sort1"] = df_merged["Ship-to-Sort1"].astype(str).str.replace(".0","", regex=False)
    if "Infor GPS Country" in df_merged.columns:
        df_merged["Infor GPS Country"] = df_merged["Infor GPS Country"].astype(str).str.replace(".0","", regex=False)

    # ===== Normalisasi Market PO Number vs Cust Ord No (zero-pad ke 10 digit) =====
    if "Cust Ord No" in df_merged.columns:
        df_merged["Cust Ord No"] = (
            df_merged["Cust Ord No"]
            .astype(str).str.strip()
            .str.replace(r"\D", "", regex=True)
            .str.zfill(10)
            .replace("0000000000", "")  # handle NaN / kosong
        )
    if "Infor Market PO Number" in df_merged.columns:
        df_merged["Infor Market PO Number"] = (
            df_merged["Infor Market PO Number"]
            .astype(str).str.strip()
            .str.replace(r"\D", "", regex=True)
            .str.zfill(10)
            .replace("0000000000", "")  # handle NaN / kosong
        )

    def safe_result(c1, c2):
        if c1 in df_merged.columns and c2 in df_merged.columns:
            return np.where(df_merged[c1] == df_merged[c2], "TRUE", "FALSE")
        return ["COLUMN MISSING"] * len(df_merged)

    df_merged["Result_Quantity"]            = safe_result("Quantity","Infor Quantity")
    df_merged["Result_Model Name"]          = safe_result("Model Name","Infor Model Name")
    df_merged["Result_Article No"]          = safe_result("Article No","Infor Article No")
    df_merged["Result_Classification Code"] = safe_result("Classification Code","Infor Classification Code")
    df_merged["Result_Delay_CRD"]           = safe_result("Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD")
    df_merged["Result_Delay_PSDD"]          = safe_result("Delay - PO PSDD Update","Infor Delay - PO PSDD Update")
    df_merged["Result_Delay_PD"]            = safe_result("Delay - PO PD Update","Infor Delay - PO PD Update")
    df_merged["Result_Lead Time"]           = safe_result("Article Lead time","Infor Lead time")
    df_merged["Result_Country"]             = safe_result("Ship-to Country","Infor Ship-to Country")
    df_merged["Result_Sort1"]               = safe_result("Ship-to-Sort1","Infor GPS Country")
    df_merged["Result_FPD"]                 = safe_result("FPD","Infor FPD")
    df_merged["Result_LPD"]                 = safe_result("LPD","Infor LPD")
    df_merged["Result_CRD"]                 = safe_result("CRD","Infor CRD")
    df_merged["Result_PSDD"]                = safe_result("PSDD","Infor PSDD")
    df_merged["Result_PODD"]                = safe_result("PODD","Infor PODD")
    df_merged["Result_PD"]                  = safe_result("PD","Infor PD")
    df_merged["Result_Market PO"]           = safe_result("Cust Ord No","Infor Market PO Number")  # <= baru
    return df_merged

DESIRED_ORDER = [
    'Client No','Site','Brand FTY Name','SO','Order Type','Order Type Description',
    'PO No.(Full)','Order Status Infor','PO No.(Short)','Merchandise Category 2','Quantity',
    'Infor Quantity','Result_Quantity','Model Name','Infor Model Name','Result_Model Name',
    'Article No','Infor Article No','Result_Article No','SAP Material','Pattern Code(Up.No.)',
    'Model No','Outsole Mold','Gender','Category 1','Category 2','Category 3','Unit Price',
    'Classification Code','Infor Classification Code','Result_Classification Code','DRC',
    'Delay/Early - Confirmation PD','Delay/Early - Confirmation CRD','Infor Delay/Early - Confirmation CRD',
    'Result_Delay_CRD','Delay - PO PSDD Update','Infor Delay - PO PSDD Update','Result_Delay_PSDD',
    'Delay - PO PD Update','Infor Delay - PO PD Update','Result_Delay_PD',
    'MDP','PDP','SDP','Article Lead time','Infor Lead time',
    'Result_Lead Time','Cust Ord No','Infor Market PO Number','Result_Market PO',  # <= baru: Cust Ord No diikuti kolom Infor & Result
    'Ship-to-Sort1','Infor GPS Country','Result_Sort1',
    'Ship-to Country','Infor Ship-to Country','Result_Country',
    'Ship to Name','Infor Shipment Method','Document Date','FPD','Infor FPD','Result_FPD','LPD','Infor LPD',
    'Result_LPD','CRD','Infor CRD','Result_CRD','PSDD','Infor PSDD','Result_PSDD',
    'FCR Date','PODD','Infor PODD','Result_PODD','PD','Infor PD','Result_PD',
    'PO Date','Actual PGI','Segment','S&P LPD','Currency','Customer PO item'
]

def reorder_columns(df, desired_order):
    existing = [c for c in desired_order if c in df.columns]
    tail = [c for c in df.columns if c not in existing]
    return df[existing + tail]

def normalize_po(x):
    if pd.isna(x):
        return ""
    x = str(x).strip()
    x = re.sub(r"\D", "", x)
    return x.zfill(10)


def process_infor_po_level(df_all):
    """
    PO-level aggregation (v5 logic):
    - Infor Quantity = SUM semua size
    - 1 PO = 1 baris
    """

    selected_columns = [
        'Order #','Order Status','Model Name','Article Number',
        'Gps Customer Number','Country/Region',
        'Customer Request Date (CRD)','Plan Date',
        'PO Statistical Delivery Date (PSDD)',
        'First Production Date','Last Production Date',
        'PODD','Production Lead Time','Class Code',
        'Delay - Confirmation','Delay - PO Del Update',
        'Delivery Delay Pd','Quantity','Shipment Method',
        'Market PO Number'  # <= baru
    ]

    missing = [c for c in selected_columns if c not in df_all.columns]
    if missing:
        st.error(f"Kolom Infor hilang: {missing}")
        return pd.DataFrame()

    df = df_all[selected_columns].copy()
    df["Order #"] = df["Order #"].apply(normalize_po)

    df_po = (
        df.groupby("Order #", as_index=False)
        .agg({
            'Order Status':'first',
            'Model Name':'first',
            'Article Number':'first',
            'Gps Customer Number':'first',
            'Country/Region':'first',
            'Customer Request Date (CRD)':'first',
            'Plan Date':'first',
            'PO Statistical Delivery Date (PSDD)':'first',
            'First Production Date':'first',
            'Last Production Date':'first',
            'PODD':'first',
            'Production Lead Time':'first',
            'Class Code':'first',
            'Delay - Confirmation':'first',
            'Delay - PO Del Update':'first',
            'Delivery Delay Pd':'first',
            'Quantity':'sum',                 # 🔥 PO-level SUM
            'Shipment Method':'first',
            'Market PO Number':'first'        # <= baru
        })
    )

    df_po.rename(columns={
        'Order Status':'Order Status Infor',
        'Model Name':'Infor Model Name',
        'Article Number':'Infor Article No',
        'Gps Customer Number':'Infor GPS Country',
        'Country/Region':'Infor Ship-to Country',
        'Customer Request Date (CRD)':'Infor CRD',
        'Plan Date':'Infor PD',
        'PO Statistical Delivery Date (PSDD)':'Infor PSDD',
        'First Production Date':'Infor FPD',
        'Last Production Date':'Infor LPD',
        'PODD':'Infor PODD',
        'Production Lead Time':'Infor Lead time',
        'Class Code':'Infor Classification Code',
        'Delay - Confirmation':'Infor Delay/Early - Confirmation CRD',
        'Delay - PO Del Update':'Infor Delay - PO PSDD Update',
        'Delivery Delay Pd':'Infor Delay - PO PD Update',
        'Quantity':'Infor Quantity',
        'Shipment Method':'Infor Shipment Method',
        'Market PO Number':'Infor Market PO Number'  # <= baru
    }, inplace=True)

    return convert_date_columns(df_po)


def build_report(df_sap, df_infor_raw):
    # === SAP ===
    df_sap2 = load_sap(df_sap)
    df_sap2["PO No.(Full)"] = df_sap2["PO No.(Full)"].apply(normalize_po)

    # === INFOR (PO LEVEL) ===
    df_infor = process_infor_po_level(df_infor_raw)
    if df_infor.empty:
        return pd.DataFrame()

    # === MERGE PO-LEVEL ===
    df_merged = df_sap2.merge(
        df_infor,
        how="left",
        left_on="PO No.(Full)",
        right_on="Order #"
    )

    df_merged = fill_missing_dates(df_merged)
    df_final  = clean_and_compare(df_merged)

    return reorder_columns(df_final, DESIRED_ORDER)


def _blank_delay_columns(df):
    out = df.copy()
    for col in DELAY_EMPTY_COLUMNS:
        if col in out.columns:
            out[col] = out[col].replace({np.nan:"", pd.NA:"", None:"", "NaN":"", "NAN":"", "NULL":"", "--":"", 0:"", 0.0:"", "0":""})
    return out

def _export_excel_styled(df, sheet_name="Report"):
    if not EXCEL_EXPORT_AVAILABLE:
        raise RuntimeError("Fitur ekspor Excel (styled) butuh 'openpyxl' (requirements.txt: openpyxl>=3.1)")
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row: cell.fill = PatternFill(fill_type=None)

        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        idx_by_name = {c.value: i+1 for i, c in enumerate(header_cells)}
        for cell in header_cells:
            col_name = str(cell.value)
            if col_name in INFOR_COLUMNS_FIXED:
                fill = PatternFill("solid", fgColor=INFOR_COLOR)
            elif col_name.startswith("Result_"):
                fill = PatternFill("solid", fgColor=RESULT_COLOR)
            else:
                fill = PatternFill("solid", fgColor=OTHER_COLOR)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=9, bold=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = PatternFill(fill_type=None)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", size=9)

        for date_col in DATE_COLUMNS_PREF:
            if date_col in idx_by_name:
                cidx = idx_by_name[date_col]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    if cell.value not in ("", None):
                        cell.number_format = DATE_FMT

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            maxlen = 0
            for cell in ws[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                maxlen = max(maxlen, len(v))
            ws.column_dimensions[col_letter].width = min(max(9, maxlen + 2), 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    bio.seek(0)
    return bio

# ================== Helpers (PO Splitter) ==================
def parse_input(text: str, split_mode: str = "auto"):
    text = text.strip()
    if not text: return []
    if split_mode == "newline": raw = text.splitlines()
    elif split_mode == "comma": raw = text.split(",")
    elif split_mode == "semicolon": raw = text.split(";")
    elif split_mode == "whitespace": raw = re.split(r"\s+", text)
    else:
        if "\n" in text:
            raw = re.split(r"[\r\n]+", text)
            split_more = []
            for line in raw:
                line = line.strip()
                if not line: continue
                if ("," in line) or (";" in line):
                    split_more.extend(re.split(r"[,;]", line))
                else:
                    split_more.append(line)
            raw = split_more
        elif ("," in text) or (";" in text):
            raw = re.split(r"[,;]", text)
        else:
            raw = re.split(r"\s+", text)
    return [x.strip() for x in raw if str(x).strip() != ""]

def normalize_items(items, keep_only_digits=False, upper_case=False, strip_prefix_suffix=False):
    normed = []
    for it in items:
        s = str(it)
        if strip_prefix_suffix: s = re.sub(r"^\W+|\W+$", "", s)
        if keep_only_digits: s = re.sub(r"\D+", "", s)
        if upper_case: s = s.upper()
        s = s.strip()
        if s != "": normed.append(s)
    return normed

def chunk_list(items, size):
    return [items[i:i+size] for i in range(0, len(items), size)]

def to_txt_bytes(lines):
    buf = io.StringIO()
    for ln in lines: buf.write(f"{ln}\n")
    return buf.getvalue().encode("utf-8")

def df_from_list(items, col_name="PO"):
    return pd.DataFrame({col_name: items})

def make_zip_bytes(chunks, basename="chunk", as_csv=True, col_name="PO"):
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, part in enumerate(chunks, start=1):
            if as_csv:
                df = df_from_list(part, col_name=col_name)
                csv_bytes = df.to_csv(index=False).encode("utf-8")
                zf.writestr(f"{basename}_{idx:02d}.csv", csv_bytes)
            else:
                zf.writestr(f"{basename}_{idx:02d}.txt", to_txt_bytes(part))
    mem.seek(0);  return mem

# ================== Tabs ==================
tab1, tab2, tab3 = st.tabs(["📦 PGD Comparison", "🧩 PO Splitter", "📈 Sales Analytics Dashboard"])

# ------------------ Tab 1: PGD Comparison ------------------
with tab1:
    st.caption("Upload 1 SAP Excel (*.xlsx) dan satu atau lebih Infor CSV (*.csv). App akan merge, cleaning, comparison, filter, dan unduhan report (Excel/CSV).")

    with st.sidebar:
        st.header("📤 Upload Files (PGD)")
        sap_file = st.file_uploader("SAP Excel (.xlsx)", type=["xlsx"], key="sap_upload")
        infor_files = st.file_uploader("Infor CSV (boleh multi-file)", type=["csv"], accept_multiple_files=True, key="infor_upload")
        st.markdown("""
**Tips:**
- SAP minimal punya `PO No.(Full)` & `Quantity`.
- Infor CSV minimal punya `PSDD`, `CRD`, dan `Line Aggregator`.
""")

    if sap_file and infor_files:
        with status_ctx("Membaca & menggabungkan file...", expanded=True) as status:
            try:
                sap_df = read_excel_file(sap_file)
                st.write("SAP dibaca:", sap_df.shape)

                infor_csv_dfs = [read_csv_file(f) for f in infor_files]
                infor_all = load_infor_from_many_csv(infor_csv_dfs)
                st.write("Total Infor (gabungan CSV):", infor_all.shape)

                if infor_all.empty:
                    _status_update(status, label="Gagal: tidak ada CSV Infor yang valid.", state="error")
                else:
                    _status_update(status, label="Sukses membaca semua file. Lanjut proses...", state="running")
                    final_df = build_report(sap_df, infor_all)

                    if final_df.empty:
                        _status_update(status, label="Gagal membuat report — periksa kolom wajib.", state="error")
                    else:
                        _status_update(status, label="Report siap! ✅", state="complete")

                        with st.sidebar.form("filters_form"):
                            st.header("🔎 Filters & Mode")
                            def uniq_vals(df, col):
                                if col in df.columns:
                                    return sorted([str(x) for x in df[col].dropna().unique().tolist()])
                                return []
                            status_opts = uniq_vals(final_df, "Order Status Infor")
                            selected_status = st.multiselect("Order Status Infor", options=status_opts, default=status_opts)
                            po_opts = uniq_vals(final_df, "PO No.(Full)")
                            selected_pos = st.multiselect("PO No.(Full)", options=po_opts, placeholder="Pilih PO (opsional)")

                            result_cols = [
                                "Result_Quantity","Result_FPD","Result_LPD","Result_CRD",
                                "Result_PSDD","Result_PODD","Result_PD","Result_Market PO"  # <= baru
                            ]
                            result_selections = {}
                            for col in result_cols:
                                opts = uniq_vals(final_df, col)
                                if opts:
                                    result_selections[col] = st.multiselect(col, options=opts, default=opts)

                            mode = st.radio("Mode tampilan data", ["Semua Kolom", "Analisis LPD PODD", "Analisis FPD PSDD"], horizontal=False)
                            submitted = st.form_submit_button("🔄 Execute / Terapkan")

                        if submitted or "df_view" in st.session_state:
                            if submitted:
                                st.session_state["selected_status"] = selected_status
                                st.session_state["selected_pos"] = selected_pos
                                st.session_state["result_selections"] = result_selections
                                st.session_state["mode"] = mode

                            selected_status = st.session_state.get("selected_status", status_opts)
                            selected_pos = st.session_state.get("selected_pos", [])
                            result_selections = st.session_state.get("result_selections", {})
                            mode = st.session_state.get("mode", "Semua Kolom")

                            df_view = final_df.copy()
                            if selected_status:
                                df_view = df_view[df_view["Order Status Infor"].astype(str).isin(selected_status)]
                            if selected_pos:
                                df_view = df_view[df_view["PO No.(Full)"].astype(str).isin(selected_pos)]
                            for col, sel in result_selections.items():
                                base_opts = uniq_vals(final_df, col)
                                if sel and set(sel) != set(base_opts):
                                    df_view = df_view[df_view[col].astype(str).isin(sel)]

                            st.session_state["df_view"] = df_view
                            st.session_state["final_df"] = final_df

                            st.subheader("🔎 Preview Hasil (After Execute)")
                            def subset(df, cols):
                                existing = [c for c in cols if c in df.columns]
                                missing = [c for c in cols if c not in df.columns]
                                if missing: st.caption(f"Kolom tidak ditemukan & di-skip: {missing}")
                                if not existing:
                                    st.warning("Tidak ada kolom yang cocok untuk mode ini.");  return pd.DataFrame()
                                return df[existing]

                            if mode == "Semua Kolom":
                                st.dataframe(df_view.head(100), use_container_width=True)
                            elif mode == "Analisis LPD PODD":
                                cols_lpd = [
                                    "PO No.(Full)","Order Status Infor","DRC",
                                    "Delay/Early - Confirmation PD","Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
                                    "Result_Delay_CRD","Delay - PO PSDD Update","Infor Delay - PO PSDD Update",
                                    "Result_Delay_PSDD","Delay - PO PD Update",
                                    "LPD","Infor LPD","Result_LPD",
                                    "PODD","Infor PODD","Result_PODD"
                                ]
                                st.dataframe(subset(df_view, cols_lpd).head(2000), use_container_width=True)
                            elif mode == "Analisis FPD PSDD":
                                cols_fpd_psdd = [
                                    "PO No.(Full)","Order Status Infor","DRC",
                                    "Delay/Early - Confirmation PD","Delay/Early - Confirmation CRD","Infor Delay/Early - Confirmation CRD",
                                    "Result_Delay_CRD","Delay - PO PSDD Update","Infor Delay - PO PSDD Update",
                                    "Result_Delay_PSDD","Delay - PO PD Update",
                                    "FPD","Infor FPD","Result_FPD",
                                    "PSDD","Infor PSDD","Result_PSDD"
                                ]
                                st.dataframe(subset(df_view, cols_fpd_psdd).head(2000), use_container_width=True)

                            st.subheader("📊 Comparison Summary (TRUE vs FALSE)")
                            existing_results = [
                                c for c in [
                                    "Result_Quantity","Result_FPD","Result_LPD","Result_CRD",
                                    "Result_PSDD","Result_PODD","Result_PD","Result_Market PO"  # <= baru
                                ] if c in df_view.columns
                            ]
                            if existing_results:
                                true_counts  = [int(df_view[c].eq("TRUE").sum()) for c in existing_results]
                                false_counts = [int(df_view[c].eq("FALSE").sum()) for c in existing_results]
                                totals       = [int(df_view[c].isin(["TRUE","FALSE"]).sum()) for c in existing_results]
                                acc = [(t / tot * 100.0) if tot > 0 else 0.0 for t, tot in zip(true_counts, totals)]

                                summary_df = pd.DataFrame({"Metric": existing_results,"TRUE": true_counts,"FALSE": false_counts,"Total (TRUE+FALSE)": totals,"TRUE %": [round(a,2) for a in acc]})
                                st.dataframe(summary_df, use_container_width=True)
                                st.bar_chart(summary_df.set_index("Metric")[["TRUE","FALSE"]])

                                false_df_sorted = pd.DataFrame({"Metric": existing_results,"FALSE": false_counts}).sort_values("FALSE", ascending=False).reset_index(drop=True)
                                st.markdown("**Distribusi FALSE (descending)**")
                                st.bar_chart(false_df_sorted.set_index("Metric")["FALSE"])
                                st.markdown("**🏆 TOP FALSE terbanyak**")
                                st.dataframe(false_df_sorted.head(min(5, len(false_df_sorted))), use_container_width=True)
                            else:
                                st.info("Kolom hasil perbandingan (Result_*) belum tersedia di data final.")

                            out_name_xlsx = f"PGD Comparison Tracking Report - {today_str_id()}.xlsx"
                            out_name_csv  = f"PGD Comparison Tracking Report - {today_str_id()}.csv"
                            df_export = _blank_delay_columns(df_view)

                            st.download_button(
                                label="⬇️ Download CSV (Filtered)",
                                data=df_export.to_csv(index=False).encode("utf-8"),
                                file_name=out_name_csv,
                                mime="text/csv",
                                use_container_width=True
                            )

                            try:
                                excel_bytes = _export_excel_styled(df_export, sheet_name="Report")
                                st.download_button(
                                    label="⬇️ Download Excel (Filtered, styled)",
                                    data=excel_bytes,
                                    file_name=out_name_xlsx,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            except Exception as ex_excel:
                                st.warning(f"Gagal membuat Excel styled: {ex_excel}")
                        else:
                            st.info("Atur filter/mode di sidebar, lalu klik **🔄 Execute / Terapkan**.")
            except Exception as e:
                _status_update(status, label="Terjadi error saat menjalankan aplikasi.", state="error")
                st.error("Detail error:")
                st.exception(e)
    else:
        st.info("Unggah file SAP & Infor di sidebar untuk mulai.")

# ------------------ Tab 2: PO Splitter ------------------
with tab2:
    st.markdown(
        """
Tempel **daftar PO** di bawah ini (boleh pisah baris, koma, titik koma, atau spasi).
App akan membagi ke potongan berisi **maksimal 5.000 PO** (atau sesuai setting).
"""
    )

    with st.expander("⚙️ Opsi Parsing & Normalisasi (opsional)", expanded=False):
        c1, c2, c3, c4, c5 = st.columns(5)
        split_mode = c1.selectbox("Mode pemisah", ["auto", "newline", "comma", "semicolon", "whitespace"])
        chunk_size = c2.number_input("Maks. PO per bagian", min_value=1, max_value=1_000_000, value=5000, step=1)
        drop_duplicates = c3.checkbox("Hapus duplikat (jaga urutan pertama)", value=False)
        keep_only_digits = c4.checkbox("Keep only digits (hapus non-digit)", value=False)
        upper_case = c5.checkbox("Upper-case (untuk alfanumerik)", value=False)
        strip_prefix_suffix = st.checkbox("Strip prefix/suffix non-alfanumerik", value=False)

    input_text = st.text_area(
        "Tempel daftar PO di sini:",
        height=220,
        placeholder="Contoh:\nPO001\nPO002\nPO003\n...\n— atau —\nPO001, PO002, PO003\n— atau —\nPO001 PO002 PO003",
        key="po_splitter_text"
    )

    process_btn = st.button("🚀 Proses & Bagi PO", key="po_splitter_btn")

    if process_btn:
        items = parse_input(input_text, split_mode=split_mode)
        original_count = len(items)

        if keep_only_digits or upper_case or strip_prefix_suffix:
            items = normalize_items(items, keep_only_digits=keep_only_digits, upper_case=upper_case, strip_prefix_suffix=strip_prefix_suffix)

        if drop_duplicates:
            items = list(dict.fromkeys(items))  # preserve order

        total = len(items)

        st.divider()
        st.subheader("📊 Ringkasan")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total input (sebelum normalisasi/duplikat)", original_count)
        c2.metric("Total setelah diproses", total)
        c3.metric("Ukuran per bagian", chunk_size)

        if total == 0:
            st.warning("Tidak ada PO terdeteksi. Cek input & opsi parsing.")
        else:
            parts = chunk_list(items, int(chunk_size))
            st.success(f"Berhasil dipecah menjadi **{len(parts)}** bagian.")

            st.markdown("### ⬇️ Unduh Semua Bagian (ZIP)")
            col_zip1, col_zip2 = st.columns(2)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_csv = f"PO_chunks_csv_{timestamp}"
            base_txt = f"PO_chunks_txt_{timestamp}"

            zip_csv = make_zip_bytes(parts, basename="PO_chunk", as_csv=True)
            col_zip1.download_button("Unduh ZIP (CSV)", data=zip_csv, file_name=f"{base_csv}.zip", mime="application/zip", use_container_width=True)

            zip_txt = make_zip_bytes(parts, basename="PO_chunk", as_csv=False)
            col_zip2.download_button("Unduh ZIP (TXT)", data=zip_txt, file_name=f"{base_txt}.zip", mime="application/zip", use_container_width=True)

            st.markdown("### 🔎 Pratinjau & Unduh per Bagian")
            colname = "PO"
            for idx, part in enumerate(parts, start=1):
                with st.expander(f"Bagian {idx} — {len(part)} PO", expanded=False):
                    df = df_from_list(part, col_name=colname)
                    st.dataframe(df, use_container_width=True, hide_index=True)

                    cdl1, cdl2 = st.columns(2)
                    csv_bytes = df.to_csv(index=False).encode("utf-8")
                    cdl1.download_button(f"Unduh Bagian {idx} (CSV)", data=csv_bytes, file_name=f"PO_chunk_{idx:02d}.csv", mime="text/csv", use_container_width=True)

                    txt_bytes = to_txt_bytes(part)
                    cdl2.download_button(f"Unduh Bagian {idx} (TXT)", data=txt_bytes, file_name=f"PO_chunk_{idx:02d}.txt", mime="text/plain", use_container_width=True)

            st.info("Tip: Jika tidak genap 5.000, bagian terakhir berisi sisa PO.")
    else:
        st.caption("Siap ketika kamu klik **Proses & Bagi PO**.")

# ------------------ Tab 3: Sales Analytics Dashboard (SAP only) ------------------
with tab3:
    st.caption("Upload **SAP Excel (.xlsx)** saja. Dashboard menampilkan KPI & grafik berbasis kolom SAP (Quantity, Unit Price, tanggal CRD/LPD/PD/PO Date, dsb).")

    sap_dash = st.file_uploader("📤 Upload SAP Excel untuk Dashboard (.xlsx)", type=["xlsx"], key="sap_dash_upload")

    if sap_dash:
        try:
            sap_df = load_sap(read_excel_file(sap_dash))   # fix Quanity->Quantity + tanggal → datetime
            # ====== Sidebar Filters ======
            with st.sidebar:
                st.header("📈 Dashboard Filters")
                # Pilihan kolom tanggal untuk time axis
                date_axis_col = st.selectbox(
                    "Pilih kolom tanggal sebagai sumbu waktu",
                    options=[c for c in ["CRD","LPD","PD","PO Date","Document Date","PSDD","FPD","PODD","Actual PGI"] if c in sap_df.columns],
                    index=0
                )

                # Rentang tanggal
                min_d = pd.to_datetime(sap_df[date_axis_col], errors="coerce").min()
                max_d = pd.to_datetime(sap_df[date_axis_col], errors="coerce").max()
                sd, ed = st.date_input(
                    "Rentang tanggal (berdasarkan pilihan sumbu waktu)",
                    value=(min_d.date() if pd.notna(min_d) else datetime(2025,1,1).date(),
                           max_d.date() if pd.notna(max_d) else datetime(2025,12,31).date())
                )

                def options(col):
                    return sorted([x for x in sap_df[col].dropna().astype(str).unique()]) if col in sap_df.columns else []

                segs   = st.multiselect("Segment", options("Segment"))
                sites  = st.multiselect("Site", options("Site"))
                brands = st.multiselect("Brand FTY Name", options("Brand FTY Name"))
                shipc  = st.multiselect("Ship-to Country", options("Ship-to Country"))
                cat2   = st.multiselect("Merchandise Category 2", options("Merchandise Category 2"))
                otype  = st.multiselect("Order Type Description", options("Order Type Description"))

            df = sap_df.copy()
            df[date_axis_col] = pd.to_datetime(df[date_axis_col], errors="coerce")
            if sd and ed:
                df = df[(df[date_axis_col] >= pd.Timestamp(sd)) & (df[date_axis_col] <= pd.Timestamp(ed) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1))]

            # Helper filter tanpa nonlocal: kembalikan df baru
            def apply_multi(df_in, col, vals):
                if vals and col in df_in.columns:
                    return df_in[df_in[col].astype(str).isin(vals)]
                return df_in
            
            # Terapkan berurutan
            for col, vals in [
                ("Segment", segs),
                ("Site", sites),
                ("Brand FTY Name", brands),
                ("Ship-to Country", shipc),
                ("Merchandise Category 2", cat2),
                ("Order Type Description", otype),
            ]:
                df = apply_multi(df, col, vals)

            # ====== Metrics ======
            st.subheader("📊 Key Metrics")
            df["__Value__"] = (pd.to_numeric(df.get("Unit Price", 0), errors="coerce").fillna(0.0) *
                               pd.to_numeric(df.get("Quantity", 0), errors="coerce").fillna(0.0))
            total_orders = df["PO No.(Full)"].nunique() if "PO No.(Full)" in df.columns else len(df)
            total_qty    = int(pd.to_numeric(df.get("Quantity", 0), errors="coerce").fillna(0).sum())
            total_value  = float(df["__Value__"].sum())
            avg_up       = float(pd.to_numeric(df.get("Unit Price", 0), errors="coerce").fillna(0).mean())

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Unique PO", f"{total_orders:,}")
            c2.metric("Total Quantity", f"{total_qty:,}")
            c3.metric("Total Value", f"${total_value:,.0f}")
            c4.metric("Avg Unit Price", f"${avg_up:,.2f}")

            # ====== Charts ======
            st.subheader("📈 Tren Bulanan")
            if df[date_axis_col].notna().any():
                df["_Month"] = df[date_axis_col].dt.to_period("M").dt.to_timestamp()
                ts = df.groupby("_Month", as_index=False).agg({"Quantity":"sum","__Value__":"sum"})
                st.line_chart(ts.set_index("_Month")[["Quantity","__Value__"]])
            else:
                st.info("Tidak ada nilai tanggal valid untuk membuat tren.")

            colA, colB = st.columns(2)

            with colA:
                st.subheader("🌍 Top Ship-to Country (Quantity)")
                if "Ship-to Country" in df.columns:
                    top_country = (df.groupby("Ship-to Country", as_index=False)["Quantity"]
                                   .sum().sort_values("Quantity", ascending=False).head(10))
                    st.bar_chart(top_country.set_index("Ship-to Country")["Quantity"])
                else:
                    st.info("Kolom 'Ship-to Country' tidak ada.")

            with colB:
                st.subheader("👟 Top Model Name (Quantity)")
                if "Model Name" in df.columns:
                    top_model = (df.groupby("Model Name", as_index=False)["Quantity"]
                                 .sum().sort_values("Quantity", ascending=False).head(10))
                    st.bar_chart(top_model.set_index("Model Name")["Quantity"])
                else:
                    st.info("Kolom 'Model Name' tidak ada.")

            st.subheader("⏱️ Distribusi Delay Code (SAP)")
            cols_delay = [c for c in ["Delay/Early - Confirmation CRD","Delay - PO PSDD Update","Delay - PO PD Update"] if c in df.columns]
            if cols_delay:
                for dc in cols_delay:
                    cnt = (df[dc].astype(str).str.upper().replace({"NAN":"","NULL":"","--":""})
                           .replace("NAT","").replace("N/A",""))
                    cnt = cnt[cnt.str.strip()!=""]
                    if len(cnt) == 0:
                        st.info(f"{dc}: tidak ada nilai.")
                        continue
                    topd = (cnt.value_counts().reset_index())
                    topd.columns = [dc, "Count"]
                    st.bar_chart(topd.set_index(dc)["Count"])
            else:
                st.info("Kolom delay SAP tidak ditemukan.")

            # ====== Download filtered data (CSV / Excel) ======
            st.subheader("⬇️ Download Dataset (Filtered)")
            csv_name = f"SalesAnalytics_SAP_Filtered_{today_str_id()}.csv"
            xlsx_name = f"SalesAnalytics_SAP_Filtered_{today_str_id()}.xlsx"
            st.download_button("Download CSV", data=df.drop(columns=["__Value__"], errors="ignore").to_csv(index=False).encode("utf-8"),
                               file_name=csv_name, mime="text/csv", use_container_width=True)

            try:
                excel_bytes = _export_excel_styled(df.drop(columns=["__Value__"], errors="ignore"), sheet_name="Sales Analytics")
                st.download_button("Download Excel (styled)", data=excel_bytes,
                                   file_name=xlsx_name,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)
            except Exception as ex_excel:
                st.warning(f"Excel styled gagal: {ex_excel}")

            # ====== Preview table ======
            st.subheader("🔎 Preview Data (max 1.000 rows)")
            st.dataframe(df.drop(columns=["__Value__"], errors="ignore").head(1000), use_container_width=True)

        except Exception as e:
            st.error("Gagal memuat/olah file SAP untuk dashboard.")
            st.exception(e)
    else:
        st.info("Unggah 1 file **SAP Excel (.xlsx)** untuk melihat dashboard.")

# ================== Debug Info ==================
with st.expander("🛠 Debug Info"):
    try:
        import platform
        st.write("Python:", sys.version)
        st.write("Platform:", platform.platform())
        st.write("Streamlit:", st.__version__)
        st.write("Pandas:", pd.__version__)
        import numpy
        st.write("NumPy:", numpy.__version__)
        st.write("openpyxl available:", EXCEL_EXPORT_AVAILABLE)
    except Exception as e:
        st.write("Failed to show debug info:", e)
